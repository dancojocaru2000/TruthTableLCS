#!/usr/bin/env python3

from logical_prop import LogicProposition, gen_table

import sys
if len(sys.argv) != 2:
	from sys import stderr
	print("Wrong number of arguments provided. Please provide the file name of the Excel file as the argument.", file=stderr)
	exit(1)

p = None
if sys.stdin.isatty() and sys.stdout.isatty():
	print(" ! is the symbol for not\n","| is the symbol for or\n","& is the symbol for and\n","> is the symbol for implies\n","= is the symbol for if and only if\n","Any upper case letter is an atomic proposition")
	p = input(" Please input a string of symbols:\n")
else:
	p = input()

prop = LogicProposition(p)
table = gen_table(prop)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from itertools import count as icount

wb = Workbook()
sheet = wb.active
sheet.title = "Truth Table"

for row_index, line in zip(icount(1), table):
	for column_name, item in zip((get_column_letter(i) for i in icount(1)), line):
		try:
			item = int(item)
		except (ValueError, TypeError):
			item = str(item)

		cell = sheet[column_name + str(row_index)]
		cell.value = item

		if row_index > 1 and column_name != 'A':
			from openpyxl.styles import colors, Font, Color
			ft = Font()
			if item == "T":
				ft = Font(color=Color(rgb='00009242'))
			elif item == "F":
				ft = Font(color=Color(rgb='00FF0000'))
			cell.font = ft

sheet.freeze_panes = sheet['B2']

if sys.argv[1].strip() == "-":
	from tempfile import NamedTemporaryFile
	with NamedTemporaryFile("wb") as tmp:
		wb.save(tmp.name)
		tmp.seek(0)
		sys.stdout.buffer.write(tmp.read())
else:
	wb.save(sys.argv[1])